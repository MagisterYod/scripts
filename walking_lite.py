import os
import openpyxl
import datetime
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from SqlAndMail import SqlAndMail


receiver_email = [
    'TCittcerDA@polymetal.ru',
    'UstsovAV@polymetal.ru',
    'kuznecovia@polymetal.ru'
]

ts = Side(border_style='thin')
td = Side(border_style='dotted')
font_stan = Font(name='Times New Roman', size=12)
font_bold = Font(name='Times New Roman', size=12, bold=True)
cell = NamedStyle(name='cell')
cell.alignment = Alignment(horizontal="center", vertical="center")
cell.border = Border(top=td, left=ts, bottom=td, right=ts)
num = 14

months = {
    '01': 'Январь',
    '02': 'Февраль',
    '03': 'Март',
    '04': 'Апрель',
    '05': 'Май',
    '06': 'Июнь',
    '07': 'Июль',
    '08': 'Август',
    '09': 'Сентябрь',
    '10': 'Октябрь',
    '11': 'Ноябрь',
    '12': 'Декабрь'
}


wb = openpyxl.load_workbook(f'C:\\asd\\Report\\LiteReportWalking.xlsx')
stop_day = SqlAndMail.cursor_data(
    query='ZSU.ASDFORM.GET_WALKING_REPORT_STOP_DAY',
    date=SqlAndMail.p_date(1)
).getvalue().fetchall()
report_day = SqlAndMail.cursor_data(
    query='ZSU.ASDFORM.GET_WALKING_REPORT_DAY',
    date=SqlAndMail.p_date(1)
).getvalue().fetchall()

ws = wb[str('Суточный отчет')]
ws['A3'] = f'за дату: {SqlAndMail.p_date(1)} г.'
for row_rd in report_day:
    if row_rd[0] == 'Время в работе за сутки, ч':
        ws['B6'] = row_rd[1]
        ws['C6'] = row_rd[2]
    if row_rd[0] == 'Объем переэкскавированной руды за сутки, м3':
        ws['B7'] = row_rd[1]
        ws['C7'] = row_rd[2]
    if row_rd[0] == 'Повторная переэкскавация за сутки, м3':
        ws['B8'] = row_rd[1]
        ws['C8'] = row_rd[2]
    if row_rd[0] == 'Время в работе за месяц, ч':
        ws['B9'] = row_rd[1]
        ws['C9'] = row_rd[2]
    if row_rd[0] == 'Объем переэкскавированной руды с начала месяца, м3':
        ws['B10'] = row_rd[1]
        ws['C10'] = row_rd[2]
    if row_rd[0] == 'Объем переэкскавированной руды с начала года, м3':
        ws['B11'] = row_rd[1]
        ws['C11'] = row_rd[2]

diff = datetime.timedelta(0)

for row_sd in stop_day:
    diff_str = str(row_sd[3] - row_sd[2]).split(':')
    cell.font = font_stan
    ws[f'A{num}'] = row_sd[1]
    ws[f'A{num}'].style = cell
    ws[f'B{num}'] = row_sd[2].strftime('%Y.%m.%d %H:%M')
    ws[f'B{num}'].style = cell
    ws[f'C{num}'] = row_sd[3].strftime('%Y.%m.%d %H:%M')
    ws[f'C{num}'].style = cell
    cell.font = font_bold
    ws[f'D{num}'] = f'{diff_str[0]}:{diff_str[1]}'
    ws[f'D{num}'].style = cell
    cell.font = font_stan
    ws[f'E{num}'] = row_sd[5]
    ws[f'E{num}'].style = cell
    diff += row_sd[3] - row_sd[2]
    num += 1

ws.merge_cells(f'A{num}:C{num}')
ws[f'A{num}'] = 'Всего простоев:'
diff = str(diff).split(':')
ws[f'D{num}'] = f'{diff[0]}:{diff[1]}'
cell.font = font_bold
cell.border = Border(top=ts, left=ts, bottom=ts, right=ts)
ws[f'A{num}'].style = cell
ws[f'B{num}'].style = cell
ws[f'C{num}'].style = cell
ws[f'D{num}'].style = cell
ws[f'E{num}'].style = cell

wb.save(f'lite_walking_{SqlAndMail.p_date(1)}.xlsx')

SqlAndMail.send_emails(
    receiver_email=receiver_email,
    body=f'Добрый день!\n\nЭто автоматическое письмо! Отвечать на него не надо.',
    subject=f'Суточный отчет по ЭШ 6/45 за {SqlAndMail.p_date(1)}',
    file_name=f'lite_walking_{SqlAndMail.p_date(1)}.xlsx'
)

os.remove(f'lite_walking_{SqlAndMail.p_date(1)}.xlsx')
print('Done')
