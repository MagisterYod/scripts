import os
import openpyxl
from typing import Callable
from SqlAndMail import SqlAndMail

receiver_email = [
    'RootSA@polymetal.ru',
    'kuznecovia@polymetal.ru'
]


def insert_vehicle(sheet, row_num, df):

    columns = [
        'D', 'E', 'F', 'G', 'I', 'J', 'L', 'M', 'N', 'O', 'R', 'S', 'T', 'U', 'X', 'Y', 'Z', 'AA', 'AB', 'AD',
        'AE', 'AF', 'AG', 'AH', 'AI', 'AJ', 'AK', 'AM', 'AN', 'AO', 'AP', 'AR', 'AS', 'AU', 'AY', 'BR', 'BS'
    ]

    for col in range(0, len(columns) - 1):
        sheet[f'{columns[col]}{row_num}'] = df[col + 1]


def insert_shovels(sheet, row_num, df, point):

    columns = [
        'D', 'E', 'F', 'G', 'H', 'J', 'L', 'M', 'N', 'O', 'R', 'S', 'T', 'V', 'W', 'X', 'Y', 'Z', 'AC', 'AD', 'AE',
        'AF', 'AG', 'AI', 'AJ', 'AK', 'AL', 'AM', 'AN', 'AO', 'AP', 'AR', 'AS', 'AT', 'AU', 'AV', 'AX', 'AY', 'BA',
        'BE'
    ]
    none_int: Callable[[None, float], float] = lambda x: x if x is not None else 0
    downtime = (float(none_int(df[14])) + float(none_int(df[15])) + float(none_int(df[16])) + float(none_int(df[17]))
                + float(none_int(df[18])) + float(none_int(df[19])) + float(none_int(df[20])) + float(none_int(df[21]))
                + float(none_int(df[22])) + float(none_int(df[23])) + float(none_int(df[24])) + float(none_int(df[25]))
                + float(none_int(df[26])) + float(none_int(df[27])) + float(none_int(df[28])) + float(none_int(df[29]))
                + float(none_int(df[30])) + float(none_int(df[31])) + float(none_int(df[32])) + float(none_int(df[33]))
                + float(none_int(df[34])) + float(none_int(df[35])) + float(none_int(df[36])) + float(none_int(df[37]))
                + float(none_int(df[38])) + float(none_int(df[39])))
    for col in range(0, len(columns)):
        sheet[f'{columns[col]}{row_num}'] = df[col + 1]
    if point == 'ir':
        sheet[f'S{row_num}'] = None if round(12 - downtime, 2) == 0 else round(12 - downtime, 2)
    if point == 'rd':
        sheet[f'T{row_num}'] = None if round(12 - downtime - float(none_int(df[12])), 2) == 0 else (
            round(12 - downtime - float(none_int(df[12])), 2))
    sheet[f'BH{row_num}'] = df[0]
    sheet[f'BI{row_num}'] = df[41]


def insert_bulls(sheet, row_num, df):

    columns = [
        'D', 'E', 'F', 'I', 'J', 'K', 'L', 'O', 'P', 'Q', 'R', 'S', 'U', 'V', 'W', 'X', 'Y', 'Z',
        'AA', 'AC', 'AD', 'AE', 'AF', 'AH', 'AI', 'AK', 'AO', 'AP', 'AS', 'AU'
    ]

    for col in range(0, len(columns) - 1):
        sheet[f'{columns[col]}{row_num}'] = df[2::][col]


def insert_allstoppagesdet(sheet, row_num, df):

    columns = ['H', 'I', 'J']

    for col in range(0, len(columns)):
        sheet[f'{columns[col]}{row_num}'] = df[2::][col]


result_query_vehicles = []
result_query_shovels = []
result_query_bulls = []
result_query_allstoppagesdet = []
result_query_wellmetervalue = []


for sh in range(1, 3):
    result_query_vehicles.append(
        [SqlAndMail.cursor_data(
            query='ZSU.EXPORTFULLREPORT.GET_VEHICLE',
            date=f'{SqlAndMail.p_date(days=1)}',
            shift=sh).getvalue().fetchall(), sh]
    )
    result_query_shovels.append(
        [SqlAndMail.cursor_data(
            query='ZSU.EXPORTFULLREPORT.GET_SHOVELS',
            date=f'{SqlAndMail.p_date(days=1)}',
            shift=sh).getvalue().fetchall(), sh])
    result_query_bulls.append(
        [SqlAndMail.cursor_data(
            query='ZSU.EXPORTFULLREPORT.GET_BULLS',
            date=f'{SqlAndMail.p_date(days=1)}',
            shift=sh).getvalue().fetchall(), sh])
    result_query_wellmetervalue.append(
        [SqlAndMail.cursor_data(
            query='ZSU.EXPORTFULLREPORT.GET_WELLMETERVALUE',
            date=f'{SqlAndMail.p_date(days=1)}',
            shift=sh).getvalue().fetchall(), sh])
result_query_allstoppagesdet.append(
    SqlAndMail.cursor_data(
        query='ZSU.GET_ALLSTOPPAGESDET',
        date=f'{SqlAndMail.p_date(days=1)}').getvalue().fetchall())
wb = openpyxl.load_workbook(f'C:\\Users\\it_kuznecovia\\Documents\\ProjectsDjango\\webASD\\po6\\static\\report.xlsx')

ws = wb[str('БелАЗы ')]
for rows in result_query_vehicles:
    if len(rows[0]) > 0:
        list_rows = rows[0]
        for row in list_rows:
            if row[0] == '30' and rows[1] == 1:
                insert_vehicle(ws, 7, row)
            if row[0] == '32' and rows[1] == 1:
                insert_vehicle(ws, 8, row)
            if row[0] == '35' and rows[1] == 1:
                insert_vehicle(ws, 9, row)
            if row[0] == '30' and rows[1] == 2:
                insert_vehicle(ws, 18, row)
            if row[0] == '32' and rows[1] == 2:
                insert_vehicle(ws, 19, row)
            if row[0] == '35' and rows[1] == 2:
                insert_vehicle(ws, 20, row)

ws = wb[str('Экскаваторы')]
for rows in result_query_shovels:
    if len(rows[0]) > 0:
        half_rows = rows[0]
        for row in half_rows:
            if row[0] == '43' and rows[1] == 1:
                insert_shovels(ws, 14, row, 'rd')
            if row[0] == '43' and rows[1] == 2:
                insert_shovels(ws, 29, row, 'rd')
            if row[0] == '183' and rows[1] == 1:
                insert_shovels(ws, 13, row, 'rd')
            if row[0] == '183' and rows[1] == 2:
                insert_shovels(ws, 28, row, 'rd')
            if row[0] == '283' and rows[1] == 1:
                insert_shovels(ws, 12, row, 'rd')
            if row[0] == '283' and rows[1] == 2:
                insert_shovels(ws, 27, row, 'rd')
            # if row[0] == '284' and rows[1] == 1:
            #     insert_shovels(ws, 15, row, 'rd')
            # if row[0] == '284' and rows[1] == 2:
            #     insert_shovels(ws, 30, row, 'rd')
            if row[0] == '285' and rows[1] == 1:
                insert_shovels(ws, 16, row, 'rd')
            if row[0] == '285' and rows[1] == 2:
                insert_shovels(ws, 31, row, 'rd')
            if row[0] == '58' and rows[1] == 1:
                insert_shovels(ws, 11, row, 'ir')
            if row[0] == '58' and rows[1] == 2:
                insert_shovels(ws, 26, row, 'ir')
            if row[0] == '45' and rows[1] == 1:
                insert_shovels(ws, 10, row, 'ir')
            if row[0] == '45' and rows[1] == 2:
                insert_shovels(ws, 25, row, 'ir')
            if row[0] == '286' and rows[1] == 1:
                insert_shovels(ws, 17, row, 'rd')
            if row[0] == '286' and rows[1] == 2:
                insert_shovels(ws, 32, row, 'rd')

ws = wb[str('Бульдозеры')]
for rows in result_query_bulls:
    if len(rows[0]) > 0:
        half_rows = rows[0]
        for row in half_rows:
            if row[0] == 9 and rows[1] == 1:
                insert_bulls(ws, 8, row)
            if row[0] == 9 and rows[1] == 2:
                insert_bulls(ws, 12, row)
            if row[0] == 186 and rows[1] == 1:
                insert_bulls(ws, 7, row)
            if row[0] == 186 and rows[1] == 2:
                insert_bulls(ws, 11, row)
            if row[0] == 4 and rows[1] == 1:
                insert_bulls(ws, 9, row)
            if row[0] == 4 and rows[1] == 2:
                insert_bulls(ws, 13, row)
            if row[0] == 15 and rows[1] == 1:
                insert_bulls(ws, 10, row)
            if row[0] == 15 and rows[1] == 2:
                insert_bulls(ws, 14, row)

ws = wb[str('Техника new')]
for rows in result_query_allstoppagesdet:
    if len(rows) > 0:
        row = rows[0]
        if row[1] == 45:
            insert_allstoppagesdet(ws, 9, row)
        if row[1] == 28:
            insert_allstoppagesdet(ws, 12, row)
        if row[1] == 31:
            insert_allstoppagesdet(ws, 13, row)
        if row[1] == 186:
            insert_allstoppagesdet(ws, 15, row)
        if row[1] == 9:
            insert_allstoppagesdet(ws, 16, row)
        if row[1] == 4:
            insert_allstoppagesdet(ws, 18, row)
        if row[1] == 15:
            insert_allstoppagesdet(ws, 19, row)
        if row[1] == 58:
            insert_allstoppagesdet(ws, 21, row)
        if row[1] == 283:
            insert_allstoppagesdet(ws, 22, row)
        if row[1] == 183:
            insert_allstoppagesdet(ws, 23, row)
        if row[1] == 43:
            insert_allstoppagesdet(ws, 24, row)
        if row[1] == 284:
            insert_allstoppagesdet(ws, 25, row)
        if row[1] == 285:
            insert_allstoppagesdet(ws, 26, row)
        if row[1] == 286:
            insert_allstoppagesdet(ws, 27, row)

ws = wb[str('Суточный рапорт')]
ws['C2'] = f'{SqlAndMail.p_date(days=1)}'
for rows in result_query_wellmetervalue:
    if len(rows[0]) > 0:
        list_rows = rows[0]
        for row in list_rows:
            if row[0] == 'Зумпф №1 СВК Ю' and rows[1] == 1:
                ws[f'AX137'] = row[1]
            if row[0] == 'Зумпф №1 СВК Ю' and rows[1] == 2:
                ws[f'CS137'] = row[1]
            if row[0] == 'Расходомер Промплощадка' and rows[1] == 1:
                ws[f'AX149'] = row[1]
            if row[0] == 'Расходомер Промплощадка' and rows[1] == 2:
                ws[f'CS149'] = row[1]

wb.save(f'report_{SqlAndMail.p_date(days=1)}.xlsx')

SqlAndMail.send_emails(
    receiver_email=receiver_email,
    body='Добрый день!\n\nЭто автоматическое рассылка! Отвечать на него не надо.\nПриятного дня!',
    subject=f'Ежесменный рапорт за дату {SqlAndMail.p_date(days=1)}',
    file_name=f'report_{SqlAndMail.p_date(days=1)}.xlsx'
)

os.remove(f'report_{SqlAndMail.p_date(days=1)}.xlsx')
print('Done Report!')
